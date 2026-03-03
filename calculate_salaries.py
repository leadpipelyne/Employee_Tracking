#!/usr/bin/env python3
"""
February 2026 Salary Calculation Script - V2
With confirmed exceptions and name mappings.
Processes each employee one by one.
"""

import json
import openpyxl
import pandas as pd
from datetime import datetime
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

# ============================================================
# CONFIGURATION
# ============================================================
MONTH = "February"
YEAR = 2026
WORKING_DAYS = 20
FULL_DAY_HOURS = 9
THRESHOLD_HOURS_PER_DAY = 8.25
MONTHLY_THRESHOLD = WORKING_DAYS * THRESHOLD_HOURS_PER_DAY  # 165 hours
MONTHLY_TOTAL_EXPECTED = WORKING_DAYS * FULL_DAY_HOURS       # 180 hours
PAID_HOLIDAYS_PER_MONTH = 2

SALARY_FILE = "/home/ubuntu/upload/February-26Salaries.xlsx"
LEAVE_FILE = "/home/ubuntu/upload/LeaveRequests.xlsx"
INSIGHTFUL_HOURS_FILE = "/home/ubuntu/insightful_feb2026_hours.json"
OUTPUT_FILE = "/home/ubuntu/February_2026_Salary_Report.xlsx"

# ============================================================
# EXCEPTIONS — Confirmed by user
# ============================================================
# Fixed salary employees: pay full, no deduction/addition
FIXED_SALARY_EXCEPTIONS = {
    "FLOID",
    "CARLA",
    "Atisha",
    "Sneha",
    "VINITA SHRIWASTAV",
    "Rakhi Devi",
    "Ronak",
}

# Full month absent: full deduction (salary = 0)
FULL_ABSENT = {
    "Jyoti",
}

# Name mappings: Salary Sheet Name -> Insightful Name (lowercase)
NAME_MAPPINGS = {
    "Ajith Kumar": "ajit kumar",
}

# ============================================================
# LOAD DATA
# ============================================================
def load_insightful_hours():
    print("=" * 80)
    print("STEP 1: Loading Insightful API Hours Data")
    print("=" * 80)
    with open(INSIGHTFUL_HOURS_FILE) as f:
        data = json.load(f)
    hours_by_name = {}
    for emp_id, info in data.items():
        name = info["name"].strip()
        hours_by_name[name.lower()] = {
            "name": name,
            "total_hours": info["total_hours"],
            "days_worked": info["days_worked"],
            "shift_count": info["shift_count"],
        }
    print(f"  Loaded hours for {len(hours_by_name)} employees from Insightful")
    return hours_by_name


def load_salary_data():
    print("\n" + "=" * 80)
    print("STEP 2: Loading Salary Data from Excel")
    print("=" * 80)
    wb = openpyxl.load_workbook(SALARY_FILE, data_only=True)
    ws = wb["Salaries"]
    salary_data = {}
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False):
        name = row[0].value
        if not name or name == "Salary For February 2026":
            continue
        salary = row[1].value
        if salary is None or salary == 0:
            continue
        currency = row[12].value if row[12].value else "INR"
        reimbursement = row[6].value if row[6].value else 0
        salary_data[name.strip()] = {
            "name": name.strip(),
            "salary": float(salary),
            "currency": currency,
            "reimbursement": float(reimbursement) if reimbursement else 0,
        }
    print(f"  Loaded salary data for {len(salary_data)} employees")
    for name, info in salary_data.items():
        print(f"    {name:30s} | Salary: {info['salary']:>12,.2f} {info['currency']}")
    return salary_data


def load_leave_data():
    print("\n" + "=" * 80)
    print("STEP 3: Loading Leave Data")
    print("=" * 80)
    wb = openpyxl.load_workbook(SALARY_FILE, data_only=True)
    ws_leaves = wb["Leaves"]
    leave_data = {}
    for row in ws_leaves.iter_rows(min_row=3, max_row=40, values_only=False):
        name = row[0].value
        if not name:
            continue
        name = name.replace('\u2060', '').strip()
        leaves_considered = row[2].value
        status = row[4].value
        leave_data[name] = {
            "name": name,
            "leaves_considered": float(leaves_considered) if leaves_considered else 0,
            "status": str(status) if status else "",
        }
    print(f"  Loaded leave data for {len(leave_data)} employees")
    for name, info in leave_data.items():
        if info["leaves_considered"] > 0:
            print(f"    {name:30s} | Leaves: {info['leaves_considered']:>5.1f} | Status: {info['status']}")
    return leave_data


# ============================================================
# NAME MATCHING
# ============================================================
def match_employee_name(salary_name, insightful_hours):
    # Check explicit name mappings first
    if salary_name in NAME_MAPPINGS:
        mapped = NAME_MAPPINGS[salary_name]
        if mapped in insightful_hours:
            return mapped

    # Direct match (case-insensitive)
    if salary_name.lower() in insightful_hours:
        return salary_name.lower()

    # First name match
    first_name = salary_name.split()[0].lower()
    matches = [k for k in insightful_hours.keys() if k.startswith(first_name)]
    if len(matches) == 1:
        return matches[0]

    # Last name match
    parts = salary_name.split()
    if len(parts) > 1:
        last_name = parts[-1].lower()
        matches = [k for k in insightful_hours.keys() if last_name in k]
        if len(matches) == 1:
            return matches[0]

    # Partial match
    for key in insightful_hours.keys():
        if first_name in key or key in salary_name.lower():
            return key

    return None


def match_leave_name(salary_name, leave_data):
    clean = salary_name.replace('\u2060', '').strip()
    # Exact match
    for leave_name, linfo in leave_data.items():
        if leave_name.lower() == clean.lower():
            return linfo
    # First name match
    first = clean.split()[0].lower()
    for leave_name, linfo in leave_data.items():
        if leave_name.split()[0].lower() == first:
            return linfo
    return None


# ============================================================
# PROCESS EACH EMPLOYEE
# ============================================================
def process_employees(salary_data, insightful_hours, leave_data):
    print("\n" + "=" * 80)
    print("STEP 4: Processing Each Employee One by One")
    print("=" * 80)

    results = []

    for salary_name, sal_info in salary_data.items():
        print(f"\n{'─' * 70}")
        print(f"PROCESSING: {salary_name}")
        print(f"{'─' * 70}")

        salary = sal_info["salary"]
        currency = sal_info["currency"]
        reimbursement = sal_info["reimbursement"]
        print(f"  Salary: {salary:,.2f} {currency}")

        # ── CHECK EXCEPTIONS ──
        if salary_name in FIXED_SALARY_EXCEPTIONS:
            print(f"  >>> EXCEPTION: Fixed Salary — Pay full, no deduction")
            results.append({
                "name": salary_name,
                "salary": salary,
                "currency": currency,
                "actual_hours": "N/A",
                "days_worked": "N/A",
                "leaves_considered": "N/A",
                "leave_hours": "N/A",
                "total_billable_hours": "N/A",
                "status": "FIXED",
                "deduction": 0,
                "addition": 0,
                "salary_after_deduction": salary,
                "reimbursement": reimbursement,
                "total_pay": round(salary + reimbursement, 2),
                "hourly_rate": "N/A",
                "insightful_match": "N/A (Exception)",
                "notes": "Fixed salary — no hour tracking",
            })
            continue

        if salary_name in FULL_ABSENT:
            print(f"  >>> EXCEPTION: Full Month Absent — Full deduction")
            results.append({
                "name": salary_name,
                "salary": salary,
                "currency": currency,
                "actual_hours": 0,
                "days_worked": 0,
                "leaves_considered": 0,
                "leave_hours": 0,
                "total_billable_hours": 0,
                "status": "FULL ABSENT",
                "deduction": salary,
                "addition": 0,
                "salary_after_deduction": 0,
                "reimbursement": reimbursement,
                "total_pay": round(reimbursement, 2),
                "hourly_rate": round(salary / MONTHLY_TOTAL_EXPECTED, 4),
                "insightful_match": "N/A (Full absent)",
                "notes": "Full month absent — full deduction",
            })
            continue

        # ── MATCH TO INSIGHTFUL ──
        insightful_key = match_employee_name(salary_name, insightful_hours)
        if insightful_key:
            actual_hours = insightful_hours[insightful_key]["total_hours"]
            days_worked = insightful_hours[insightful_key]["days_worked"]
            insightful_name = insightful_hours[insightful_key]["name"]
            print(f"  Insightful Match: {insightful_name}")
            print(f"  Actual Hours (API): {actual_hours:.2f}")
            print(f"  Days Worked: {days_worked}")
            match_label = f"Yes ({insightful_name})"
        else:
            actual_hours = 0
            days_worked = 0
            print(f"  Insightful Match: NOT FOUND (0 hours)")
            match_label = "No"

        # ── MATCH TO LEAVE DATA ──
        leave_info = match_leave_name(salary_name, leave_data)
        if leave_info:
            leaves_considered = leave_info["leaves_considered"]
            print(f"  Leave Match: {leave_info['name']} | Leaves: {leaves_considered}")
        else:
            leaves_considered = 0
            print(f"  Leave Match: NOT FOUND (0 leaves)")

        # ── CALCULATE ──
        leave_hours = leaves_considered * FULL_DAY_HOURS
        total_billable_hours = actual_hours + leave_hours
        hourly_rate = salary / MONTHLY_TOTAL_EXPECTED if MONTHLY_TOTAL_EXPECTED > 0 else 0

        print(f"  Leave Hours Credit: {leave_hours:.2f} ({leaves_considered} × {FULL_DAY_HOURS})")
        print(f"  Total Billable Hours: {total_billable_hours:.2f}")
        print(f"  Threshold: {MONTHLY_THRESHOLD:.2f} | Expected: {MONTHLY_TOTAL_EXPECTED:.2f}")

        deduction = 0
        addition = 0
        final_status = ""
        notes = ""

        if total_billable_hours < MONTHLY_THRESHOLD:
            shortfall = MONTHLY_TOTAL_EXPECTED - total_billable_hours
            deduction = shortfall * hourly_rate
            final_status = "DEDUCT"
            notes = f"Short by {shortfall:.2f} hrs from {MONTHLY_TOTAL_EXPECTED}"
            print(f"  >>> DEDUCT: Shortfall {shortfall:.2f} hrs × {hourly_rate:.2f} = {deduction:,.2f}")
        elif total_billable_hours > MONTHLY_TOTAL_EXPECTED:
            excess = total_billable_hours - MONTHLY_TOTAL_EXPECTED
            addition = excess * hourly_rate
            final_status = "ADDITION"
            notes = f"Overtime {excess:.2f} hrs above {MONTHLY_TOTAL_EXPECTED}"
            print(f"  >>> ADDITION: Excess {excess:.2f} hrs × {hourly_rate:.2f} = {addition:,.2f}")
        else:
            final_status = "OK"
            notes = "Within threshold, no adjustment"
            print(f"  >>> OK: No deduction or addition")

        salary_after = salary - deduction + addition
        total_pay = salary_after + reimbursement

        print(f"  Final: {salary:,.2f} - {deduction:,.2f} + {addition:,.2f} = {salary_after:,.2f}")

        results.append({
            "name": salary_name,
            "salary": salary,
            "currency": currency,
            "actual_hours": round(actual_hours, 2),
            "days_worked": days_worked,
            "leaves_considered": leaves_considered,
            "leave_hours": round(leave_hours, 2),
            "total_billable_hours": round(total_billable_hours, 2),
            "status": final_status,
            "deduction": round(deduction, 2),
            "addition": round(addition, 2),
            "salary_after_deduction": round(salary_after, 2),
            "reimbursement": reimbursement,
            "total_pay": round(total_pay, 2),
            "hourly_rate": round(hourly_rate, 4),
            "insightful_match": match_label,
            "notes": notes,
        })

    return results


# ============================================================
# GENERATE EXCEL REPORT
# ============================================================
def generate_report(results):
    print("\n" + "=" * 80)
    print("STEP 5: Generating Final Salary Report")
    print("=" * 80)

    wb = openpyxl.Workbook()

    # ── Styles ──
    hdr_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    deduct_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    add_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ok_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fixed_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    absent_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    num_fmt = '#,##0.02'

    # ═══════════════════════════════════════════════════════
    # SHEET 1: February 2026 Salaries
    # ═══════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "February 2026 Salaries"

    # Title
    ws.merge_cells("A1:Q1")
    c = ws["A1"]
    c.value = f"Salary Report — {MONTH} {YEAR}"
    c.font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:Q2")
    c = ws["A2"]
    c.value = (f"Working Days: {WORKING_DAYS}  |  Threshold: {MONTHLY_THRESHOLD} hrs  "
               f"|  Total Expected: {MONTHLY_TOTAL_EXPECTED} hrs  |  Paid Holidays: {PAID_HOLIDAYS_PER_MONTH} days/month")
    c.font = Font(name="Calibri", size=10, italic=True)
    c.alignment = Alignment(horizontal="center")

    headers = [
        "Employee Name", "Currency", "Monthly Salary",
        "Actual Hours\n(Insightful)", "Days\nWorked",
        "Leaves\nConsidered", "Leave Hours\nCredit",
        "Total Billable\nHours", "Status",
        "Deduction", "Addition",
        "Salary After\nAdjustment", "Reimbursement", "Total Pay",
        "Hourly Rate", "Insightful\nMatch", "Notes",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin

    numeric_cols = {3, 10, 11, 12, 13, 14}
    hour_cols = {4, 7, 8, 15}

    for ri, r in enumerate(results, 5):
        row_data = [
            r["name"], r["currency"], r["salary"],
            r["actual_hours"], r["days_worked"],
            r["leaves_considered"], r["leave_hours"],
            r["total_billable_hours"], r["status"],
            r["deduction"], r["addition"],
            r["salary_after_deduction"], r["reimbursement"], r["total_pay"],
            r["hourly_rate"], r["insightful_match"], r["notes"],
        ]

        status_fills = {
            "DEDUCT": deduct_fill, "ADDITION": add_fill,
            "OK": ok_fill, "FIXED": fixed_fill, "FULL ABSENT": absent_fill,
        }
        row_fill = status_fills.get(r["status"], ok_fill)

        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center" if ci != 1 and ci != 17 else "left")

            if ci in numeric_cols and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
            elif ci in hour_cols and isinstance(val, (int, float)):
                cell.number_format = '0.00'

            if ci == 9:
                cell.fill = row_fill
                cell.font = Font(bold=True)
            elif ci == 10 and isinstance(val, (int, float)) and val > 0:
                cell.fill = deduct_fill
                cell.font = Font(color="9C0006", bold=True)
            elif ci == 11 and isinstance(val, (int, float)) and val > 0:
                cell.fill = add_fill
                cell.font = Font(color="006100", bold=True)

    # Totals row
    tot_row = len(results) + 6
    ws.cell(row=tot_row, column=1, value="TOTALS").font = Font(bold=True, size=12)
    for col in [3, 10, 11, 12, 13, 14]:
        letter = openpyxl.utils.get_column_letter(col)
        cell = ws.cell(row=tot_row, column=col,
                       value=f"=SUM({letter}5:{letter}{tot_row - 1})")
        cell.font = Font(bold=True)
        cell.number_format = '#,##0.00'
        cell.border = thin

    widths = [28, 10, 15, 14, 8, 10, 12, 14, 13, 14, 14, 16, 14, 14, 12, 22, 35]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    # ═══════════════════════════════════════════════════════
    # SHEET 2: Comparison with Original
    # ═══════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Comparison with Original")

    wb_orig = openpyxl.load_workbook(SALARY_FILE, data_only=True)
    ws_orig = wb_orig["Salaries"]
    original = {}
    for row in ws_orig.iter_rows(min_row=3, max_row=ws_orig.max_row, values_only=False):
        name = row[0].value
        if name and name != "Salary For February 2026":
            original[name.strip()] = {
                "salary": row[1].value or 0,
                "deduction": row[3].value or 0,
                "addition": row[4].value or 0,
                "hours": row[8].value or 0,
            }

    comp_headers = [
        "Employee Name", "Salary",
        "Orig Deduction", "Orig Addition", "Orig Hours",
        "New Hours (API)", "New Leaves", "New Total Hrs",
        "New Deduction", "New Addition",
        "Diff Deduction", "Diff Addition", "Notes",
    ]
    for ci, h in enumerate(comp_headers, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin

    for ri, r in enumerate(results, 2):
        orig = original.get(r["name"], {"salary": 0, "deduction": 0, "addition": 0, "hours": 0})
        orig_ded = orig["deduction"] or 0
        orig_add = orig["addition"] or 0
        new_ded = r["deduction"] if isinstance(r["deduction"], (int, float)) else 0
        new_add = r["addition"] if isinstance(r["addition"], (int, float)) else 0
        diff_ded = new_ded - orig_ded
        diff_add = new_add - orig_add

        act_hrs = r["actual_hours"] if isinstance(r["actual_hours"], (int, float)) else ""
        leaves = r["leaves_considered"] if isinstance(r["leaves_considered"], (int, float)) else ""
        tot_hrs = r["total_billable_hours"] if isinstance(r["total_billable_hours"], (int, float)) else ""

        notes = r.get("notes", "")
        if abs(diff_ded) > 500 or abs(diff_add) > 500:
            notes = "⚠ SIGNIFICANT CHANGE  |  " + notes

        data = [
            r["name"], r["salary"],
            orig_ded, orig_add, orig["hours"],
            act_hrs, leaves, tot_hrs,
            new_ded, new_add,
            round(diff_ded, 2), round(diff_add, 2), notes,
        ]
        for ci, val in enumerate(data, 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.border = thin
            if ci in {2, 3, 4, 5, 6, 8, 9, 10, 11, 12} and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
            if ci == 11 and isinstance(val, (int, float)) and abs(val) > 500:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            if ci == 12 and isinstance(val, (int, float)) and abs(val) > 500:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for i, w in enumerate([28, 14, 14, 14, 12, 14, 10, 14, 14, 14, 14, 14, 40], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # ═══════════════════════════════════════════════════════
    # SHEET 3: Exceptions & Config
    # ═══════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Exceptions & Config")

    ws3.cell(row=1, column=1, value="Configuration").font = Font(bold=True, size=13)
    config_items = [
        ("Month", f"{MONTH} {YEAR}"),
        ("Working Days", WORKING_DAYS),
        ("Full Day Hours", FULL_DAY_HOURS),
        ("Threshold Hours/Day", THRESHOLD_HOURS_PER_DAY),
        ("Monthly Threshold", MONTHLY_THRESHOLD),
        ("Monthly Total Expected", MONTHLY_TOTAL_EXPECTED),
        ("Paid Holidays/Month", PAID_HOLIDAYS_PER_MONTH),
    ]
    for ri, (k, v) in enumerate(config_items, 3):
        ws3.cell(row=ri, column=1, value=k).font = Font(bold=True)
        ws3.cell(row=ri, column=2, value=v)

    ws3.cell(row=12, column=1, value="Fixed Salary Exceptions").font = Font(bold=True, size=13)
    ws3.cell(row=13, column=1, value="Employee Name").font = Font(bold=True)
    ws3.cell(row=13, column=2, value="Reason").font = Font(bold=True)
    for ri, name in enumerate(sorted(FIXED_SALARY_EXCEPTIONS), 14):
        ws3.cell(row=ri, column=1, value=name)
        ws3.cell(row=ri, column=2, value="Fixed salary — no hour-based deduction")

    r_start = 14 + len(FIXED_SALARY_EXCEPTIONS) + 2
    ws3.cell(row=r_start, column=1, value="Full Absent Exceptions").font = Font(bold=True, size=13)
    for ri, name in enumerate(sorted(FULL_ABSENT), r_start + 1):
        ws3.cell(row=ri, column=1, value=name)
        ws3.cell(row=ri, column=2, value="Full month absent — full deduction")

    r_start2 = r_start + len(FULL_ABSENT) + 3
    ws3.cell(row=r_start2, column=1, value="Name Mappings").font = Font(bold=True, size=13)
    ws3.cell(row=r_start2 + 1, column=1, value="Salary Sheet Name").font = Font(bold=True)
    ws3.cell(row=r_start2 + 1, column=2, value="Insightful Name").font = Font(bold=True)
    for ri, (k, v) in enumerate(NAME_MAPPINGS.items(), r_start2 + 2):
        ws3.cell(row=ri, column=1, value=k)
        ws3.cell(row=ri, column=2, value=v)

    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 45

    # ═══════════════════════════════════════════════════════
    # SHEET 4: Deduction Logic
    # ═══════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Calculation Logic")
    logic_lines = [
        "SALARY DEDUCTION / ADDITION LOGIC",
        "",
        "1. Each employee is expected to work 9 hours per day.",
        "2. A 45-minute lunch relief reduces the threshold to 8 hours 15 minutes per day.",
        f"3. Monthly Threshold = Working Days ({WORKING_DAYS}) × 8.25 = {MONTHLY_THRESHOLD} hours.",
        f"4. Monthly Total Expected = Working Days ({WORKING_DAYS}) × 9 = {MONTHLY_TOTAL_EXPECTED} hours.",
        "",
        "5. Total Billable Hours = Actual Hours (from Insightful) + (Approved Leaves × 9).",
        "",
        "6. IF Total Billable Hours < Monthly Threshold (165):",
        "       → DEDUCT",
        f"       → Shortfall = {MONTHLY_TOTAL_EXPECTED} − Total Billable Hours",
        "       → Deduction = Shortfall × (Salary / 180)",
        "       NOTE: Deduction is from the FULL 9-hour day, not from 8h15m threshold.",
        "",
        f"7. IF Total Billable Hours > Monthly Total Expected ({MONTHLY_TOTAL_EXPECTED}):",
        "       → ADDITION (Overtime)",
        f"       → Excess = Total Billable Hours − {MONTHLY_TOTAL_EXPECTED}",
        "       → Addition = Excess × (Salary / 180)",
        "",
        f"8. IF Total Billable Hours is between {MONTHLY_THRESHOLD} and {MONTHLY_TOTAL_EXPECTED}:",
        "       → OK — No deduction or addition.",
        "",
        "9. Fixed salary employees are paid in full regardless of hours.",
        "10. Full absent employees receive full deduction (salary = 0).",
    ]
    for ri, line in enumerate(logic_lines, 1):
        cell = ws4.cell(row=ri, column=1, value=line)
        if ri == 1:
            cell.font = Font(bold=True, size=14)
        elif line.startswith(("1.", "2.", "3.", "4.", "5.", "6.", "7.", "8.", "9.", "10.")):
            cell.font = Font(bold=True)
    ws4.column_dimensions["A"].width = 80

    # ── Save ──
    wb.save(OUTPUT_FILE)
    print(f"\n  Report saved to: {OUTPUT_FILE}")
    return OUTPUT_FILE


# ============================================================
# MAIN
# ============================================================
def main():
    print("=" * 80)
    print(f"  MANGOEYES — AUTOMATED SALARY CALCULATION")
    print(f"  {MONTH} {YEAR}")
    print(f"  Working Days: {WORKING_DAYS}")
    print(f"  Threshold: {MONTHLY_THRESHOLD} hrs | Expected: {MONTHLY_TOTAL_EXPECTED} hrs")
    print("=" * 80)

    insightful_hours = load_insightful_hours()
    salary_data = load_salary_data()
    leave_data = load_leave_data()
    results = process_employees(salary_data, insightful_hours, leave_data)
    output = generate_report(results)

    # ── Summary ──
    print("\n" + "=" * 80)
    print("FINAL SUMMARY")
    print("=" * 80)

    # Only sum numeric deductions/additions
    total_ded = sum(r["deduction"] for r in results if isinstance(r["deduction"], (int, float)))
    total_add = sum(r["addition"] for r in results if isinstance(r["addition"], (int, float)))
    total_sal = sum(r["salary"] for r in results)
    total_final = sum(r["salary_after_deduction"] for r in results if isinstance(r["salary_after_deduction"], (int, float)))

    counts = defaultdict(int)
    for r in results:
        counts[r["status"]] += 1

    print(f"  Total Employees:       {len(results)}")
    print(f"  Deductions:            {counts['DEDUCT']}")
    print(f"  Additions (Overtime):  {counts['ADDITION']}")
    print(f"  OK (No Change):        {counts['OK']}")
    print(f"  Fixed Salary:          {counts['FIXED']}")
    print(f"  Full Absent:           {counts['FULL ABSENT']}")
    print(f"  Total Deductions:      {total_ded:,.2f}")
    print(f"  Total Additions:       {total_add:,.2f}")
    print(f"  Total Gross Salary:    {total_sal:,.2f}")
    print(f"  Total Net Salary:      {total_final:,.2f}")
    print(f"\n  Report: {output}")

    with open("/home/ubuntu/salary_results_v2.json", "w") as f:
        json.dump(results, f, indent=2, default=str)


if __name__ == "__main__":
    main()
